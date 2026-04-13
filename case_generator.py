"""
测试用例生成器 - 根据检索到的知识生成测试用例
支持批量生成大量测试用例
"""

import json
import re
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Union
from langchain_core.documents import Document

# Excel支持
try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, Border, Side

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class TestCaseGenerator:
    """测试用例生成器"""

    def __init__(self, llm_provider, output_dir: Path):
        self.llm_provider = llm_provider
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def _build_cases_prompt(
            self,
            *,
            query: str,
            context: str,
            batch_size: int,
            examples: str = "",
            existing_names: List[str] = None
    ) -> str:
        """构建更结构化、更短、更稳定的用例生成提示词。"""
        existing_names = existing_names or []

        examples_section = f"\n[参考示例]\n{examples}\n" if examples else ""

        existing_section = ""
        if existing_names:
            # 只给少量已存在名称，避免prompt膨胀
            names_str = "、".join([f'"{n}"' for n in existing_names[:20]])
            existing_section = f"\n[已存在用例名(勿重复)]\n{names_str}\n"

        # 结构化：任务 -> 约束 -> Schema -> 示例 -> 输入
        # 重点：减少长篇“角色设定”，明确“只输出JSON数组”。
        return (
            "[任务]\n"
            f"基于需求与知识库内容，生成 {batch_size} 个测试用例（覆盖：主流程/边界/异常/关联）。\n\n"
            "[输出要求]\n"
            "- 只输出：一个标准JSON数组。不要markdown、不要解释、不要额外文本。\n"
            "- JSON必须可被 json.loads() 直接解析；必须以 `[` 开头、以 `]` 结尾。\n"
            "- 不要生成与“已存在用例名”重复的 name。\n"
            "- 文本字段中不要出现真实换行符；如需换行，用字面 `\\n`。\n\n"
            "[JSON Schema]\n"
            "[\n"
            "  {\n"
            '    "name": "用例名称(仅step_id=1填写)",\n'
            '    "step_id": 1,\n'
            '    "step": "1. ...\\\\n2. ...",\n'
            '    "precondition": "前置条件",\n'
            '    "priority": "高|中|低",\n'
            '    "expected": "预期结果"\n'
            "  },\n"
            "  {\n"
            '    "name": "",\n'
            '    "step_id": 2,\n'
            '    "step": "1. ...\\\\n2. ...",\n'
            '    "precondition": "",\n'
            '    "priority": "",\n'
            '    "expected": "预期结果"\n'
            "  }\n"
            "]\n\n"
            "[字段规则]\n"
            "- 每个元素必须包含：name, step_id, step, precondition, priority, expected 六个键。\n"
            "- 同一用例允许多行：step_id=1 行填写 name/priority；step_id>1 行 name/priority 置空字符串。\n"
            "- step_id 必须是数字(1,2,3...)。\n\n"
            f"{examples_section}"
            f"{existing_section}"
            "[输入]\n"
            f"需求：{query}\n\n"
            "知识库内容：\n"
            f"{context}\n\n"
            "[现在开始输出JSON数组]\n"
        )

    def generate(
            self,
            query: str,
            context_docs: List[Document],
            num_cases: int = 10,
            batch_size: int = 10,
            max_retries: int = 2,
            examples: str = "",
            use_streaming: bool = True
    ) -> List[Dict]:
        """
        生成测试用例（支持大批量）

        Args:
            query: 查询/需求
            context_docs: 检索到的上下文文档
            num_cases: 需要生成的测试用例数量
            batch_size: 每批生成的用例数量
            max_retries: 每批最大重试次数
            examples: 用例示例，用于指导生成格式和风格
            use_streaming: 是否使用流式接收模式（默认True，可获取完整JSON）

        Returns:
            生成的测试用例内容
        """
        print(f"\n开始生成测试用例（目标: {num_cases} 个，每批 {batch_size} 个）...")

        all_test_cases = []
        existing_names_set = set()  # 记录已生成的用例名，避免重复（O(1) membership）

        # 持续生成直到达到目标数量
        batch_num = 0
        # 防止LLM持续失败导致死循环：按“目标批次数 * 每批重试”上限兜底
        max_total_batches = max(1, (num_cases + batch_size - 1) // batch_size) * max(1, max_retries) * 3
        while len(all_test_cases) < num_cases:
            if batch_num >= max_total_batches:
                print(f"  已达到最大批次数上限({max_total_batches})，提前停止生成。")
                break
            current_batch_size = min(batch_size, num_cases - len(all_test_cases))
            start_idx = len(all_test_cases) + 1

            print(
                f"  生成第 {batch_num + 1} 批 (目标: {current_batch_size} 个用例，已获取: {len(all_test_cases)}/{num_cases})...")

            batch_cases = []
            retry_count = 0

            # 选择生成方法：流式优先，回退到普通模式
            if use_streaming:
                generate_method = self._generate_batch_streaming
            else:
                generate_method = self._generate_batch

            # 重试直到获得足够用例或达到最大重试次数
            while len(batch_cases) < current_batch_size and retry_count < max_retries:
                # 请求数量等于目标数量
                request_size = current_batch_size - len(batch_cases)
                current_start_idx = start_idx + len(batch_cases)

                temp_cases = generate_method(
                    query, context_docs,
                    start_idx=current_start_idx,
                    batch_size=request_size,
                    examples=examples,
                    existing_names=list(existing_names_set)
                )

                if temp_cases:
                    # 过滤掉与已有用例重复的（根据name判断）
                    new_cases = []
                    # 子步骤必须能归属到最近主步骤；如果本批前面没有主步骤，
                    # 允许归属到历史已保留主步骤（existing_names_set 非空）。
                    has_main_anchor = len(existing_names_set) > 0
                    for tc in temp_cases:
                        name = tc.get('name', '').strip()
                        step_id = tc.get('step_id')
                        step_text = (tc.get('step') or '').strip()

                        # 主步骤（step_id=1）按 name 去重
                        if name and name not in existing_names_set:
                            new_cases.append(tc)
                            existing_names_set.add(name)
                            has_main_anchor = True
                        elif name:
                            print(f"    过滤掉重复用例: {name}")
                        # 子步骤（step_id>1 且 name 为空）也要保留，否则会丢失 step_id=2/3...
                        elif step_id and step_id > 1 and step_text and has_main_anchor:
                            new_cases.append(tc)
                        elif step_id and step_id > 1 and step_text:
                            print(f"    过滤掉孤立子步骤: step_id={step_id}, step='{step_text[:30]}'")

                    batch_cases.extend(new_cases)
                    print(
                        f"    第 {retry_count + 1} 次尝试: 请求{request_size}个, 获取 {len(temp_cases)} 个, 去重后 {len(new_cases)} 个")
                else:
                    print(f"    第 {retry_count + 1} 次尝试: 获取 0 个用例")

                retry_count += 1

            if batch_cases:
                all_test_cases.extend(batch_cases)
                print(f"  本批实际获取: {len(batch_cases)} 个用例")
            else:
                print(f"  第 {batch_num + 1} 批生成失败")

            print(f"  当前累计: {len(all_test_cases)}/{num_cases} 个测试用例")
            batch_num += 1

        if not all_test_cases:
            return []

        # 去除重复的用例（根据name去重，保留第一个）
        unique_cases = []
        seen_names = set()
        for tc in all_test_cases:
            name = tc.get('name', '').strip()
            step_id = tc.get('step_id')
            step_text = (tc.get('step') or '').strip()
            if name and name not in seen_names:
                seen_names.add(name)
                unique_cases.append(tc)
            elif name:
                print(f"  去除重复用例: {name}")
            # 保留多步骤子行（name 为空是预期行为）
            elif step_id and step_id > 1 and step_text:
                unique_cases.append(tc)

        # 二次防护：去除无法归属到最近主步骤(step_id=1且name非空)的子步骤行
        unique_cases = self._drop_orphan_sub_steps(unique_cases)

        print(f"\n总计生成: {len(all_test_cases)} 个测试用例，去重后: {len(unique_cases)} 个")

        return unique_cases

    def _generate_batch(
            self,
            query: str,
            context_docs: List[Document],
            start_idx: int = 1,
            batch_size: int = 10,
            examples: str = "",
            existing_names: List[str] = None
    ) -> List[Dict]:
        """生成一批测试用例"""
        if existing_names is None:
            existing_names = []

        print(f"  _generate_batch: 开始生成 {batch_size} 个用例")
        # 构建上下文
        context = "\n\n".join([
            f"文档{i + 1}:\n{doc.page_content[:1000]}"  # 限制每个文档长度
            for i, doc in enumerate(context_docs[:5])  # 最多5个文档
        ])

        prompt = self._build_cases_prompt(
            query=query,
            context=context,
            batch_size=batch_size,
            examples=examples,
            existing_names=existing_names
        )

        # 重试机制：最多重试次数
        max_retries = 3
        retry_count = 0

        while retry_count < max_retries:
            try:
                response = self.llm_provider.chat(prompt)

                # 打印完整原始响应用于调试
                print(f"  LLM原始响应:\n{response[:2000]}")

                # 检测JSON是否完整
                if not self._is_json_complete(response):
                    retry_count += 1
                    if retry_count < max_retries:
                        print(f"  检测到JSON可能被截断（第 {retry_count} 次重试），重新请求...")
                        # 添加提醒到prompt中
                        reminder = "\n\n[警告：请务必返回完整的JSON数组，以 [ 开头以 ] 结尾，不要截断！]"
                        prompt = prompt + reminder
                        continue
                    else:
                        print(f"  JSON截断重试 {max_retries} 次仍失败，尝试解析不完整JSON...")

                # 解析JSON响应
                cases = self._parse_json_response(response, expected_count=batch_size)

                # 如果解析结果为空且还有重试次数，也重试
                if not cases and retry_count < max_retries - 1:
                    retry_count += 1
                    print(f"  解析结果为空（第 {retry_count} 次重试），重新请求...")
                    reminder = "\n\n[警告：请务必返回完整的JSON数组！]"
                    prompt = prompt + reminder
                    continue

                print(f"  解析得到 {len(cases)} 条用例")
                for i, c in enumerate(cases[:5]):
                    print(
                        f"    {i + 1}. name='{c.get('name', '')}', step_id={c.get('step_id')}, step='{c.get('step', '')}', precondition='{c.get('precondition', '')}', priority='{c.get('priority', '')}', expected='{c.get('expected', '')}'")
                return cases
            except Exception as e:
                import traceback
                print(f"  批次生成失败: {e}")
                traceback.print_exc()
                retry_count += 1
                if retry_count < max_retries:
                    print(f"  重试中（第 {retry_count} 次）...")

        return []

    def _generate_batch_streaming(
            self,
            query: str,
            context_docs: List[Document],
            start_idx: int = 1,
            batch_size: int = 10,
            examples: str = "",
            existing_names: List[str] = None
    ) -> List[Dict]:
        """使用流式接收生成一批测试用例（确保获取完整JSON）

        Args:
            query: 查询/需求
            context_docs: 检索到的上下文文档
            start_idx: 用例起始索引
            batch_size: 需要生成的用例数量
            examples: 用例示例
            existing_names: 已存在的用例名列表

        Returns:
            解析后的用例列表
        """
        if existing_names is None:
            existing_names = []

        print(f"  _generate_batch_streaming: 开始生成 {batch_size} 个用例（流式接收模式）...")

        # 检查是否支持流式接收
        if not hasattr(self.llm_provider, 'chat_streaming'):
            print("  警告: 当前LLM提供商不支持流式接收，回退到普通模式")
            return self._generate_batch(
                query, context_docs, start_idx, batch_size, examples, existing_names
            )

        # ============================================================
        # 原有的流式模式 prompt 构建代码（已注释）
        # ============================================================
        # # 构建上下文
        # context = "\n\n".join([
        #     f"文档{i + 1}:\n{doc.page_content[:1000]}"
        #     for i, doc in enumerate(context_docs[:5])
        # ])
        #
        # # 构建示例部分
        # examples_section = ""
        # if examples:
        #     examples_section = f"\n参考示例:\n{examples}\n"
        #
        # # 构建已存在用例的提示
        # existing_section = ""
        # if existing_names:
        #     names_str = "、".join([f'"{n}"' for n in existing_names[:20]])
        #     existing_section = f"\n注意：以下用例名已经存在，请勿生成重复的：{names_str}\n"
        #
        # # 系统提示词 - 明确要求输出标准JSON数组格式
        # system_prompt = """你是一个专业的高级测试工程师，擅长根据需求文档生成高质量的测试用例。
        #
        # 重要要求：
        # 1. 每个测试用例必须包含6个字段：name、step_id、step、precondition、priority、expected
        # 2. step字段中多个步骤用"\\n"分隔（如："1. 打开登录页面\\n2. 输入用户名"）
        # 3. 步骤中的实际换行符用\\n表示，不要有真正的换行符
        # 4. 同一个用例的多个步骤组，step_id=1填写name和priority，step_id>1的name和priority填空
        # 5. 不要生成与已有用例名重复的测试用例
        # 6. 【最重要】请直接返回标准JSON数组格式，不要使用markdown代码块包裹，不要使用函数调用格式"""
        #
        # # 用户提示词 - 明确要求JSON数组
        # user_prompt = f"""基于以下需求和知识库内容，生成 {batch_size} 个测试用例。
        #
        # 需求: {query}
        # {examples_section}
        # 知识库内容:
        # {context}
        # {existing_section}
        #
        # 【重要】请直接返回标准JSON数组格式，格式如下（不要使用```包裹，不要用函数调用格式）：
        # [
        #   {{"name":"用例名称1","step_id":1,"step":"1. 步骤1\\n2. 步骤2","precondition":"前置条件","priority":"高","expected":"预期结果"}},
        #   {{"name":"用例名称2","step_id":1,"step":"1. 步骤1\\n2. 步骤2","precondition":"前置条件","priority":"中","expected":"预期结果"}}
        # ]
        #
        # 请直接输出JSON数组："""
        # ============================================================

        # 构建上下文（与普通模式保持一致）
        context = "\n\n".join([
            f"文档{i + 1}:\n{doc.page_content[:1000]}"
            for i, doc in enumerate(context_docs[:5])
        ])

        prompt = self._build_cases_prompt(
            query=query,
            context=context,
            batch_size=batch_size,
            examples=examples,
            existing_names=existing_names
        )

        # 打印发送给 LLM 的原始请求
        print(f"\n{'='*80}")
        print(f"【发送给 LLM 的原始请求】")
        print(f"{'='*80}")
        print(f"\n[prompt]:\n{prompt}")
        print(f"\n{'='*80}\n")

        # 流式接收（使用统一后的 prompt）
        print(f"  开始流式接收LLM响应...")
        full_content = ""
        chunk_count = 0

        try:
            for chunk in self.llm_provider.chat_streaming(
                message=prompt,
                system_prompt=None
            ):
                full_content += chunk
                chunk_count += 1

            print(f"  流式接收完成，共 {chunk_count} 个 chunks，内容长度: {len(full_content)}")

            # 打印原始响应
            print(f"  LLM原始响应（前500字符）:\n{full_content[:500]}")

            # 解析JSON响应
            cases = self._parse_json_response_streaming(full_content, expected_count=batch_size)

            print(f"  解析得到 {len(cases)} 条用例")
            if cases:
                for i, c in enumerate(cases[:5]):
                    print(
                        f"    {i + 1}. name='{c.get('name', '')}', step_id={c.get('step_id')}, step='{c.get('step', '')}', precondition='{c.get('precondition', '')}', priority='{c.get('priority', '')}', expected='{c.get('expected', '')}'")
            return cases

        except Exception as e:
            import traceback
            print(f"  流式批次生成失败: {e}")
            traceback.print_exc()
            # 回退到普通模式
            return self._generate_batch(
                query, context_docs, start_idx, batch_size, examples, existing_names
            )

    def _parse_json_response_streaming(self, response: str, expected_count: int = 10) -> List[Dict]:
        """解析流式接收的JSON响应

        Args:
            response: 完整的响应内容
            expected_count: 期望的用例数量

        Returns:
            解析后的用例列表
        """
        if not response or not response.strip():
            return []

        # 清理响应，去除markdown代码块标记
        cleaned_response = response.strip()
        if cleaned_response.startswith('```'):
            lines = cleaned_response.split('\n')
            cleaned_response = '\n'.join(lines[1:])
            if cleaned_response.endswith('```'):
                cleaned_response = cleaned_response[:-3]
            cleaned_response = cleaned_response.strip()

        # 处理LLM返回实际换行符的情况
        cleaned_response = self._fix_json_newlines(cleaned_response)

        # 尝试直接解析
        try:
            cases = json.loads(cleaned_response)
            if isinstance(cases, list):
                for case in cases:
                    if 'step' in case and isinstance(case['step'], str):
                        case['step'] = case['step'].replace('\n', '\\n')
                print(f"  直接解析成功，获取 {len(cases)} 个用例")
                return cases
            else:
                return [cases]
        except json.JSONDecodeError as e:
            print(f"  JSON解析失败: {e}")

        # JSON解析失败，尝试正则提取完整对象
        cases = self._fix_truncated_json(cleaned_response, expected_count)
        if cases:
            return self._fix_case_newlines(cases)

        return []

    def _parse_json_response(self, response: str, expected_count: int = 5) -> List[Dict]:
        """解析JSON响应"""
        # 清理响应，去除markdown代码块标记
        cleaned_response = response.strip()
        if cleaned_response.startswith('```'):
            # 去除 ```json 或 ``` 标记
            lines = cleaned_response.split('\n')
            cleaned_response = '\n'.join(lines[1:])
            if cleaned_response.endswith('```'):
                cleaned_response = cleaned_response[:-3]
            cleaned_response = cleaned_response.strip()

        # 处理LLM返回实际换行符的情况：将字符串值中的实际换行符替换为\\n
        cleaned_response = self._fix_json_newlines(cleaned_response)

        # 尝试直接解析
        try:
            cases = json.loads(cleaned_response)
            if isinstance(cases, list):
                # 修复：step字段中的真实换行符替换为\\n字符串
                for case in cases:
                    if 'step' in case and isinstance(case['step'], str):
                        case['step'] = case['step'].replace('\n', '\\n')
                print(f"  成功解析JSON，获取 {len(cases)} 个用例")
                return cases
            else:
                return [cases]
        except Exception as e:
            print(f"  JSON解析失败: {e}")
            # JSON解析失败，尝试修复截断的JSON
            cases = self._fix_truncated_json(cleaned_response, expected_count)
            if cases:
                # 修复：所有文本字段中的真实换行符替换为\n字符串
                cases = self._fix_case_newlines(cases)
                return cases

        try:
            # 尝试提取JSON数组
            match = re.search(r'\[.*\]', cleaned_response, re.DOTALL)
            if match:
                json_str = match.group()
                cases = json.loads(json_str)
                if isinstance(cases, list):
                    print(f"  成功解析JSON，获取 {len(cases)} 个用例")
                    # 修复：所有文本字段中的真实换行符替换为\n字符串
                    return self._fix_case_newlines(cases)
                else:
                    cases = [cases]
                    return self._fix_case_newlines(cases)

            # 尝试直接解析
            cases = json.loads(cleaned_response)
            if isinstance(cases, list):
                return self._fix_case_newlines(cases)
            else:
                return self._fix_case_newlines([cases])
        except Exception as e:
            print(f"  JSON解析失败: {e}")
            # JSON解析失败，尝试修复截断的JSON
            cases = self._fix_truncated_json(cleaned_response, expected_count)
            return self._fix_case_newlines(cases) if cases else []

    def _fix_json_newlines(self, json_str: str) -> str:
        """修复JSON中未转义的换行符"""
        result = []
        in_string = False
        escape_next = False
        i = 0
        while i < len(json_str):
            char = json_str[i]
            if escape_next:
                result.append(char)
                escape_next = False
            elif char == '\\':
                result.append(char)
                escape_next = True
            elif char == '"':
                result.append(char)
                in_string = not in_string
            elif char == '\n' and in_string:
                # 在字符串内的换行符需要转义
                result.append('\\n')
            elif char == '\r':
                # 忽略回车符
                pass
            else:
                result.append(char)
            i += 1
        return ''.join(result)

    def _is_json_complete(self, json_str: str) -> bool:
        """检测JSON字符串是否完整（可正常解析）

        Args:
            json_str: JSON字符串

        Returns:
            True if JSON is complete and valid, False otherwise
        """
        if not json_str or not json_str.strip():
            return False

        # 1. 检查括号平衡
        if json_str.count('[') != json_str.count(']'):
            return False
        if json_str.count('{') != json_str.count('}'):
            return False

        # 2. 检查引号平衡（排除转义引号）
        in_string = False
        escape_next = False
        for char in json_str:
            if escape_next:
                escape_next = False
                continue
            if char == '\\':
                escape_next = True
                continue
            if char == '"':
                in_string = not in_string

        if in_string:  # 引号未闭合
            return False

        # 3. 尝试解析验证
        try:
            json.loads(json_str)
            return True
        except:
            return False

    def _fix_case_newlines(self, cases: List[Dict]) -> List[Dict]:
        """修复用例列表中所有文本字段的换行符（真实换行符转为\n字符串）"""
        text_fields = ['name', 'step', 'precondition', 'expected', 'priority']
        for case in cases:
            for field in text_fields:
                if field in case and isinstance(case[field], str):
                    case[field] = case[field].replace('\n', '\\n')
        return cases

    def _fix_truncated_json(self, json_str: str, expected_count: int) -> List[Dict]:
        """尝试修复截断的JSON"""
        cases = []

        # 方法1：找到最后一个完整的对象
        try:
            array_start = json_str.find('[')
            if array_start >= 0:
                # 找到最后一个 }
                last_brace = json_str.rfind('}')
                if last_brace > array_start:
                    fixed_json = json_str[array_start:last_brace + 1] + ']'
                    # 修复未关闭的字符串
                    fixed_json = self._close_unclosed_strings(fixed_json)
                    fixed_cases = json.loads(fixed_json)
                    if isinstance(fixed_cases, list) and len(fixed_cases) > 0:
                        print(f"  修复截断JSON成功，获取 {len(fixed_cases)} 个用例")
                        return fixed_cases
        except Exception as e:
            print(f"  修复尝试1失败: {e}")

        # 方法2：正则提取每个完整对象
        try:
            case_blocks = re.findall(r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}', json_str, re.DOTALL)
            for block in case_blocks[:expected_count]:
                case = self._extract_case_from_block(block)
                # 只有当用例有有效 name 或 step 时才添加
                if case and (case.get('name') or case.get('step')):
                    cases.append(case)
            if cases:
                print(f"  正则提取获取 {len(cases)} 个用例")
                return cases
        except Exception as e:
            print(f"  修复尝试2失败: {e}")

        return cases

    def _close_unclosed_strings(self, json_str: str) -> str:
        """关闭JSON中未关闭的字符串"""
        result = []
        in_string = False
        escape_next = False
        i = 0
        while i < len(json_str):
            char = json_str[i]
            if escape_next:
                result.append(char)
                escape_next = False
            elif char == '\\':
                result.append(char)
                escape_next = True
            elif char == '"':
                result.append(char)
                in_string = not in_string
            else:
                # 不能丢弃字符串中的内容；这里只做“是否在字符串内”的跟踪
                result.append(char)
            i += 1

        # 如果字符串未关闭，添加闭合引号和括号
        if in_string:
            result.append('"')
        return ''.join(result)

    def _extract_case_from_block(self, block: str) -> Dict:
        """从JSON块中提取用例信息"""
        case = {}
        try:
            # 尝试直接解析这个块
            obj = json.loads(block)
            return obj
        except:
            # 回退到正则提取
            name_match = re.search(r'["\']?name["\']?\s*:\s*["\']([^"\']*)["\']', block)
            if name_match:
                case['name'] = name_match.group(1).strip()
            step_id_match = re.search(r'["\']?step_id["\']?\s*:\s*(\d+)', block)
            if step_id_match:
                case['step_id'] = int(step_id_match.group(1))
            step_match = re.search(r'["\']?step["\']?\s*:\s*["\']([^"\']*)["\']', block, re.DOTALL)
            if step_match:
                case['step'] = step_match.group(1).replace('\n', '\\n')
            precondition_match = re.search(r'["\']?precondition["\']?\s*:\s*["\']([^"\']*)["\']', block)
            if precondition_match:
                case['precondition'] = precondition_match.group(1).strip()
            priority_match = re.search(r'["\']?priority["\']?\s*:\s*["\']([^"\']*)["\']', block)
            if priority_match:
                case['priority'] = priority_match.group(1).strip()
            expected_match = re.search(r'["\']?expected["\']?\s*:\s*["\']([^"\']*)["\']', block)
            if expected_match:
                case['expected'] = expected_match.group(1).strip()
        return case

    def _drop_orphan_sub_steps(self, cases: List[Dict]) -> List[Dict]:
        """移除无法归属到最近主步骤的子步骤行。"""
        if not cases:
            return cases

        filtered = []
        has_main_anchor = False
        orphan_count = 0

        for tc in cases:
            name = (tc.get('name') or '').strip()
            step_id = tc.get('step_id')
            step_text = (tc.get('step') or '').strip()

            if step_id == 1 and name:
                filtered.append(tc)
                has_main_anchor = True
                continue

            if step_id and step_id > 1 and step_text:
                if has_main_anchor:
                    filtered.append(tc)
                else:
                    orphan_count += 1
                continue

            # 其他场景保持原样，避免误删异常数据，交给后续流程处理
            filtered.append(tc)

        if orphan_count:
            print(f"  过滤掉 {orphan_count} 条无法归属主步骤的子步骤")
        return filtered

    def _merge_multi_step_cases(self, cases: List[Dict]) -> List[Dict]:
        """将 step_id>1 且 name 为空的行归并到最近一个非空 name 的用例下。

        说明：
        - prompt 明确要求 step_id>1 行的 name 为空，因此不能按 name 分组；
        - 这里用“最近一次出现的 step_id=1 且 name 非空”的用例作为归属。
        - 输出仍保持 step_id>1 行 name 为空（便于导出模板写入规则）。
        """
        if not cases:
            return cases

        merged = []
        current_name = ""
        current_main_seen = False

        for tc in cases:
            name = (tc.get('name') or '').strip()
            step_id = tc.get('step_id', 1)

            if step_id == 1 and name:
                current_name = name
                current_main_seen = True
                merged.append(tc)
                continue

            if not name and step_id and step_id > 1 and current_main_seen and current_name:
                # 归并到当前用例：保持name为空，但补齐缺失字段，避免Excel写入时报KeyError/None混乱
                merged.append({
                    'name': '',
                    'step_id': step_id,
                    'step': tc.get('step', ''),
                    'precondition': tc.get('precondition', ''),
                    'priority': tc.get('priority', ''),
                    'expected': tc.get('expected', '')
                })
                continue

            # 兜底：无法归属的行原样保留（例如第一行就是step_id>1，或LLM没按约定输出）
            merged.append(tc)

        print(f"  合并多步骤(按最近name归属): {len(cases)} -> {len(merged)} 行")
        return merged

    def _format_cases(self, cases: List[Dict]) -> str:
        """格式化测试用例为Markdown"""
        md_content = f"# 集成测试用例\n\n"
        md_content += f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
        md_content += f"共 {len(cases)} 个测试用例\n\n"
        md_content += "---\n\n"

        for i, tc in enumerate(cases, 1):
            md_content += f"## 测试用例 {i}: {tc.get('name', '未命名')}\n\n"

            precond = tc.get('precondition', '无')
            if precond and precond != '无':
                md_content += f"**前置条件**: {precond}\n\n"

            md_content += "**测试步骤**:\n"
            step_text = tc.get('step', '')
            if isinstance(step_text, str) and step_text:
                for line in step_text.replace('\\n', '\n').split('\n'):
                    line = line.strip()
                    if line:
                        md_content += f"{line}\n"
            else:
                md_content += "1. （无）\n"

            md_content += f"\n**预期结果**: {tc.get('expected', '')}\n\n"

            test_data = tc.get('test_data', '')
            if test_data and test_data != '无':
                md_content += f"**测试数据**: {test_data}\n\n"

            md_content += "---\n\n"

        return md_content

    def save_to_file(self, content: str, filename: str = None) -> Path:
        """保存测试用例到文件"""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"test_case_{timestamp}.md"

        filepath = self.output_dir / filename
        filepath.write_text(content, encoding='utf-8')
        print(f"\n测试用例已保存到: {filepath}")
        return filepath

    def save_to_excel(self, cases: Union[List[Dict], str], filename: str = None) -> Path:
        """保存测试用例到Excel文件

        Args:
            cases: 测试用例列表（Dict）或Markdown内容（str）
            filename: 文件名

        Returns:
            保存的文件路径
        """
        if not EXCEL_AVAILABLE:
            print("警告: openpyxl未安装，保存为Markdown格式")
            if isinstance(cases, str):
                return self.save_to_file(cases)
            return self.save_to_file(self._format_cases(cases))

        # 如果传入的是字符串（Markdown格式），先解析为用例列表
        if isinstance(cases, str):
            cases = self._parse_markdown_cases(cases)

        # 过滤无效用例：必须要有step字段
        valid_cases = [c for c in cases if c.get('step') and c.get('step').strip()]
        if len(valid_cases) < len(cases):
            print(f"  过滤掉 {len(cases) - len(valid_cases)} 个无效用例")
        cases = valid_cases

        # 合并多步骤用例：同一个name的step_id>1的行合并到step_id=1的行
        cases = self._merge_multi_step_cases(cases)

        # 保存前统计用例数量
        step1_cases = [c for c in cases if c.get('step_id') == 1]
        unique_names = set(c.get('name', '') for c in cases if c.get('name', '').strip())
        print(
            f"  保存到Excel前: 共 {len(cases)} 行, step_id=1用例: {len(step1_cases)} 个, 唯一name: {len(unique_names)} 个")

        if not cases:
            print("警告: 没有测试用例可保存")
            return self.save_to_file("无测试用例")

        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"test_case_{timestamp}.xlsx"

        filepath = self.output_dir / filename

        # 读取模板文件
        template_path = Path(__file__).parent / "templates" / "测试用例模板.xlsx"
        try:
            wb = openpyxl.load_workbook(str(template_path))
            ws = wb.active
            # 删除模板原有的空数据行，只保留表头
        except Exception as e:
            print(f"警告: 无法读取模板文件 {template_path}: {e}")
            # 创建新工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "测试用例"

        # 找到数据开始行（跳过表头）
        start_row = 2

        # 计算用例编号：仅在 step_id=1 且 name 非空时递增
        case_number = 0
        for tc in cases:
            name = (tc.get('name') or '').strip()
            step_id = tc.get('step_id')
            if step_id == 1 and name:
                case_number += 1
                tc['_case_number'] = case_number
            else:
                tc['_case_number'] = None

        # 写入测试用例数据
        for idx, tc in enumerate(cases):
            row_idx = start_row + idx

            # 获取字段值
            step_id = tc.get('step_id')  # 默认None，不是1
            name = tc.get('name', '')
            precondition = tc.get('precondition', '')
            priority = tc.get('priority', '')
            step = tc.get('step', '')
            expected = tc.get('expected', '')
            case_num = tc.get('_case_number')

            # 修复：将所有文本字段中的字面 \n 字符串替换为真正的换行符
            # LLM 返回的文本中包含的是 \n 字符串（两个字符），需要转换
            name = name.replace('\\n', '\n') if isinstance(name, str) else name
            precondition = precondition.replace('\\n', '\n') if isinstance(precondition, str) else precondition
            step = step.replace('\\n', '\n') if isinstance(step, str) else step
            expected = expected.replace('\\n', '\n') if isinstance(expected, str) else expected

            # 编号（第1列）- 只在 step_id=1 时填写（表示用例序号）
            if case_num is not None:
                ws.cell(row=row_idx, column=1, value=case_num)
            else:
                ws.cell(row=row_idx, column=1, value='')

            # 标题（第4列）- 只在 step_id=1 时填写
            if step_id == 1:
                ws.cell(row=row_idx, column=4, value=name)
            else:
                ws.cell(row=row_idx, column=4, value='')

            # 前置条件（第6列）- 每一行填写
            ws.cell(row=row_idx, column=6, value=precondition)

            # 优先级（第7列）- 只在 step_id=1 时填写
            if step_id == 1:
                ws.cell(row=row_idx, column=7, value=priority)
            else:
                ws.cell(row=row_idx, column=7, value='')

            # 步骤ID（第8列）- 只有 step_id 存在时才填写
            if step_id is not None:
                ws.cell(row=row_idx, column=8, value=step_id)
            else:
                ws.cell(row=row_idx, column=8, value='')

            # 步骤（第9列）- 用换行符分隔多个步骤
            ws.cell(row=row_idx, column=9, value=step)

            # 期望结果（第10列）
            ws.cell(row=row_idx, column=10, value=expected)

            # 设置行高以适应内容（根据换行符数量调整）
            max_lines = max(step.count('\n'), expected.count('\n')) + 1
            ws.row_dimensions[row_idx].height = max(20, max_lines * 20)

        # 删除最后一条用例之后的空行（保留表头）
        last_case_row = len(cases) + 1
        while ws.max_row > last_case_row:
            ws.delete_rows(ws.max_row)

        # 保存文件
        wb.save(filepath)
        print(f"\n测试用例已保存到: {filepath}")
        return filepath

    def _parse_markdown_cases(self, markdown_text: str) -> List[Dict]:
        """解析Markdown格式的测试用例为字典列表"""
        cases = []

        # 提取每个测试用例块（从 ## 测试用例 到下一个 ## 或结尾）
        pattern = r'##\s*测试用例\s*\d+:\s*(.+?)(?=\n##\s*测试用例|\Z)'
        matches = re.findall(pattern, markdown_text, re.DOTALL)

        for match in matches:
            case = {'name': '', 'steps': [], 'expected': '', 'priority': ''}

            # 按行处理
            lines = match.strip().split('\n')
            if not lines:
                continue

            # 第一行是标题
            case['name'] = lines[0].strip()

            # 后续行提取steps和expected
            current_field = None
            for line in lines[1:]:
                line_stripped = line.strip()

                # 跳过空行和分隔符
                if not line_stripped or line_stripped == '---':
                    continue

                # 检测字段标记（如 "**测试步骤**:" 或 "**预期结果**: ..."）
                # 需要检测是否同时包含字段名和内容（用冒号分隔）
                if '测试步骤' in line_stripped:
                    current_field = 'steps'
                    # 如果冒号后面还有内容，添加到steps
                    if ':' in line_stripped:
                        parts = line_stripped.split(':', 1)
                        if len(parts) > 1 and parts[1].strip():
                            step = parts[1].strip().replace('**', '').strip()
                            if step:
                                case['steps'].append(step)
                    continue
                elif '预期结果' in line_stripped:
                    current_field = 'expected'
                    # 如果冒号后面还有内容，添加到expected
                    if ':' in line_stripped:
                        parts = line_stripped.split(':', 1)
                        if len(parts) > 1 and parts[1].strip():
                            exp = parts[1].strip().replace('**', '').strip()
                            if exp:
                                case['expected'] = exp
                    continue

                # 提取内容
                if current_field == 'steps':
                    # 去掉序号如 "1. " 或 "1. "
                    step = re.sub(r'^\d+[.、]\s*', '', line_stripped)
                    # 去掉markdown加粗标记
                    step = step.replace('**', '').strip()
                    if step:
                        case['steps'].append(step)

                elif current_field == 'expected':
                    # 去掉markdown加粗标记
                    exp = line_stripped.replace('**', '').strip()
                    if exp:
                        if case['expected']:
                            case['expected'] += '\n' + exp
                        else:
                            case['expected'] = exp

            cases.append(case)

        return cases

